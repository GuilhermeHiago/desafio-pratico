from django.contrib import admin
from django.contrib.auth import admin as auth_admin
from sqlalchemy import column, null

from .forms import UserChangeForm, UserCreationForm
from .models import User

from django.contrib.auth import get_user_model
from django.http import HttpResponse
import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

User = get_user_model()

def export_to_xlsx(modeladmin, request, queryset):

    opts = modeladmin.model._meta

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(opts.verbose_name)

    wb = Workbook()
    # wb.active
    ws = wb.active#wb.create_sheet('Django API Users Data')
    ws.title = 'Django API Users Data'
    ws.column_dimensions
    
    # filter all fields by not many and by not blank
    fields = [field for field in opts.get_fields() if not field.many_to_many and not field.one_to_many and not field.blank]
    print("\n\nresposta:", fields)

    # Write a first row with header information
    for i in range(1, len(fields)): 
        ws.cell(1, i, fields[i].name)

    # Write data rows
    for obj in queryset:
        data_row = []
        row = 2
        for field in fields:
            value = getattr(obj, field.name)
            if isinstance(value, datetime.datetime):
                value = value.strftime('%d/%m/%Y')
            data_row.append(value)

        for i in range(1, len(data_row)): 
            ws.cell(row, i, value=data_row[i])
            ws.column_dimensions[get_column_letter(i)].auto_size = True
        # writer.writerow(data_row)

    wb.save(response)

    return response

@admin.register(User)
class UserAdmin(auth_admin.UserAdmin):
    form = UserChangeForm
    add_form = UserCreationForm
    model = User
    fieldsets = auth_admin.UserAdmin.fieldsets + (
        (None, {"fields": ("birth_date",)}),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'birth_date', 'password1', 'password2')}
        ),
    )

    actions = [export_to_xlsx]
